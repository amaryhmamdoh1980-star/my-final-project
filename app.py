import os
import requests
import base64
import json
import re
import hashlib
import io
from flask import Flask, render_template, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text

app = Flask(__name__)

DB_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres.yimsexytrswzamnslgcd:MaAm%40036355972@aws-0-ap-northeast-1.pooler.supabase.com:6543/postgres')
app.config['SQLALCHEMY_DATABASE_URI'] = DB_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024
db_obj = SQLAlchemy(app)

class DB:
    def execute(self, query, *args):
        formatted_query = query
        params = {}
        for i, arg in enumerate(args):
            placeholder = f"val{i}"
            formatted_query = formatted_query.replace("?", f":{placeholder}", 1)
            params[placeholder] = arg
        result = db_obj.session.execute(text(formatted_query), params)
        db_obj.session.commit()
        if query.strip().upper().startswith("SELECT"):
            return [dict(row._mapping) for row in result]
        return None

db = DB()
API_KEY    = os.environ.get("GOOGLE_API_KEY")
YT_API_KEY = os.environ.get("YOUTUBE_API_KEY")

MAP_WORDS   = ["מפה", "מפת", "خريطة", "خريطه"]
IMAGE_WORDS = ["צייר", "תמונה", "תראה לי", "תצלום", "איור",
               "صورة", "صوره", "أرني", "ارسم", "رسم", "صور", "اعرض"]
VIDEO_WORDS = [
    "סרטון", "וידאו", "סרט", "הראה לי סרטון", "תראה לי סרטון",
    "תן לי סרטון", "תראה לי וידאו", "רוצה לראות", "תן לי וידאו",
    "فيديو", "مقطع فيديو", "مقطع", "فلم", "شاهد",
    "أرني فيديو", "أرني مقطع", "اعرض لي فيديو", "اعرض فيديو",
    "أريد مشاهدة", "بث", "تشغيل",
]

def detect_lang(text):
    if re.search(r'[\u0600-\u06FF]', text):
        return 'ar'
    if re.search(r'[\u0590-\u05FF]', text):
        return 'he'
    return 'en'

HE_TO_EN = {
    "תראה לי את": "", "תראה לי": "",
    "צייר לי": "", "צייר": "",
    "תמונה של": "", "תמונה": "",
    "מפה של": "", "מפת": "", "מפה": "",
    "תצלום של": "", "תצלום": "",
    "איור של": "", "איור": "",
    "סרטון על": "", "סרטון של": "", "סרטון": "",
    "וידאו על": "", "וידאו של": "", "וידאו": "",
    "סרט על": "", "סרט": "",
    "הראה לי סרטון על": "", "תראה לי סרטון על": "",
    "תן לי סרטון על": "", "רוצה לראות": "",
    "של": "", "את": "", "לי": "", "אז": "", "על": "",
    "ישראל": "Israel", "ירושלים": "Jerusalem", "תל אביב": "Tel Aviv",
    "חיפה": "Haifa", "אילת": "Eilat", "הנגב": "Negev",
    "הגליל": "Galilee", "הכרמל": "Carmel",
    "ים המלח": "Dead Sea", "כנרת": "Sea of Galilee",
    "ים התיכון": "Mediterranean Sea", "ים סוף": "Red Sea",
    "הר": "mountain", "מדבר": "desert", "יער": "forest",
    "חוף": "beach", "נהר": "river", "אגם": "lake",
    "שיש": "marble", "מרמור": "marble",
    "סלע דולומיט": "dolomite", "דולומיט": "dolomite",
    "אבן גיר": "limestone", "גיר": "limestone",
    "בזלת": "basalt", "גרניט": "granite",
    "סלע": "rock", "אבן": "stone", "מינרל": "mineral",
    "קריסטל": "quartz", "קוורץ": "quartz",
    "זהב": "gold", "כסף": "silver", "נחושת": "copper",
    "ריוליט": "rhyolite", "אנדזיט": "andesite",
    "גברו": "gabbro", "דיאבז": "diabase",
    "פרידוטיט": "peridotite", "טוף": "tuff",
    "פומיס": "pumice", "אובסידיאן": "obsidian",
    "בזלת עמודי": "columnar basalt", "לבה": "lava",
    "אבן חול": "sandstone", "אבן חולית": "sandstone",
    "צור": "flint", "חלמיש": "flint",
    "שיסט חרסיתי": "shale", "חרסית": "clay",
    "גיר אלמוגים": "coral limestone", "קונגלומרט": "conglomerate",
    "ברקציה": "breccia", "חוואר": "marl",
    "גנייס": "gneiss", "שיסט": "schist",
    "קוורציט": "quartzite", "פילייט": "phyllite",
    "הורנפלס": "hornfels", "מיגמטיט": "migmatite",
    "פיריט": "pyrite", "קלציט": "calcite", "פלדספר": "feldspar",
    "מיקה": "mica", "מוסקוביט": "muscovite",
    "ביוטיט": "biotite", "גבס": "gypsum", "הליט": "halite",
    "מגנטיט": "magnetite", "המטיט": "hematite",
    "לימוניט": "limonite", "גתיט": "goethite",
    "מלכיט": "malachite", "אזוריט": "azurite",
    "גלנה": "galena", "ספלריט": "sphalerite",
    "כלקופיריט": "chalcopyrite",
    "אוליבין": "olivine", "פירוקסן": "pyroxene",
    "אמפיבול": "amphibole", "הורנבלנד": "hornblende",
    "גרנט": "garnet", "אפידוט": "epidote",
    "טלק": "talc", "כלוריט": "chlorite",
    "זרקון": "zircon", "אפטיט": "apatite",
    "טורמלין": "tourmaline", "פלואוריט": "fluorite",
    "יהלום": "diamond", "אודם": "ruby",
    "ספיר": "sapphire", "אמרלד": "emerald",
    "אמתיסט": "amethyst", "טופז": "topaz",
    "אוניקס": "onyx", "אופל": "opal",
    "ירקן": "jade", "פנינה": "pearl",
    "לפיס לזולי": "lapis lazuli",
    "פחם": "coal", "זרחן": "phosphorite", "גופרית": "sulfur",
    "שמש": "sun", "ירח": "moon", "אש": "fire",
    "דינוזאור": "dinosaur", "פיל": "elephant",
    "אריה": "lion", "נשר": "eagle", "כריש": "shark",
    "עץ": "tree", "פרח": "flower", "ורד": "rose",
    "צרפת": "France", "גרמניה": "Germany", "ספרד": "Spain",
    "איטליה": "Italy", "יוון": "Greece", "מצרים": "Egypt",
    "ירדן": "Jordan", "סוריה": "Syria", "לבנון": "Lebanon",
    "ארצות הברית": "United States", "אמריקה": "United States",
    "רוסיה": "Russia", "סין": "China", "יפן": "Japan",
    "הודו": "India", "ברזיל": "Brazil", "אוסטרליה": "Australia",
    "קנדה": "Canada", "בריטניה": "United Kingdom", "טורקיה": "Turkey",
}

AR_TO_EN = {
    "أرني": "", "ارسم لي": "", "ارسم": "",
    "صورة من": "", "صورة لـ": "", "صورة": "", "صوره": "",
    "رسم": "", "رسمة": "", "اعرض لي": "", "اعرض": "",
    "أريد مشاهدة": "", "أرني فيديو عن": "", "أرني مقطع عن": "",
    "فيديو عن": "", "فيديو": "", "مقطع فيديو عن": "", "مقطع فيديو": "",
    "مقطع عن": "", "مقطع": "", "فلم عن": "", "فلم": "",
    "شاهد": "", "بث": "", "تشغيل": "",
    "خريطة": "map", "خريطه": "map",
    "إسرائيل": "Israel", "فلسطين": "Palestine",
    "القدس": "Jerusalem", "تل أبيب": "Tel Aviv",
    "حيفا": "Haifa", "إيلات": "Eilat",
    "النقب": "Negev", "الجليل": "Galilee",
    "البحر الميت": "Dead Sea", "بحيرة طبريا": "Sea of Galilee",
    "البحر المتوسط": "Mediterranean Sea", "البحر الأحمر": "Red Sea",
    "جبل": "mountain", "صحراء": "desert",
    "غابة": "forest", "شاطئ": "beach",
    "نهر": "river", "بحيرة": "lake",
    "بازلت": "basalt", "بازالت": "basalt",
    "جرانيت": "granite", "ريوليت": "rhyolite",
    "أنديزيت": "andesite", "غابرو": "gabbro",
    "بيريدوتيت": "peridotite",
    "توف بركاني": "tuff", "توف": "tuff",
    "بيوميس": "pumice", "خفاف": "pumice",
    "أوبسيديان": "obsidian", "زجاج بركاني": "obsidian",
    "حمم": "lava", "لافا": "lava",
    "بازلت عمودي": "columnar basalt",
    "رخام": "marble", "مرمر": "marble",
    "حجر جيري": "limestone", "كلس": "limestone",
    "دولوميت": "dolomite", "حجر رملي": "sandstone",
    "صوان": "flint", "طفلة": "shale",
    "طين": "clay", "مارل": "marl",
    "كونغلوميرات": "conglomerate", "بريشيا": "breccia",
    "فحم": "coal",
    "نيس": "gneiss", "شيست": "schist",
    "كوارتزيت": "quartzite", "فيليت": "phyllite",
    "هورنفلس": "hornfels",
    "كوارتز": "quartz", "معدن": "mineral",
    "صخرة": "rock", "حجر": "stone",
    "بيريت": "pyrite", "كالسيت": "calcite",
    "فيلدسبار": "feldspar", "ميكا": "mica",
    "جبس": "gypsum", "ملح صخري": "halite",
    "مغنتيت": "magnetite", "هيماتيت": "hematite",
    "مالاكيت": "malachite", "أزوريت": "azurite",
    "غالينا": "galena", "كالكوبيريت": "chalcopyrite",
    "أوليفين": "olivine", "بيروكسين": "pyroxene",
    "أمفيبول": "amphibole", "غارنيت": "garnet",
    "تالك": "talc", "فلوريت": "fluorite",
    "توربالين": "tourmaline", "زيركون": "zircon",
    "ماسة": "diamond", "الماسة": "diamond",
    "ياقوت": "ruby", "زمرد": "emerald",
    "أميثيست": "amethyst", "توباز": "topaz",
    "عقيق": "agate", "لؤلؤ": "pearl",
    "لازورد": "lapis lazuli",
    "ذهب": "gold", "فضة": "silver",
    "نحاس": "copper", "حديد": "iron",
    "شمس": "sun", "قمر": "moon", "نار": "fire",
    "بركان": "volcano", "ديناصور": "dinosaur",
    "فيل": "elephant", "أسد": "lion",
    "شجرة": "tree", "وردة": "rose",
    "مصر": "Egypt", "الأردن": "Jordan", "سوريا": "Syria",
    "لبنان": "Lebanon", "السعودية": "Saudi Arabia",
    "فرنسا": "France", "ألمانيا": "Germany", "إسبانيا": "Spain",
    "إيطاليا": "Italy", "اليونان": "Greece",
    "الولايات المتحدة": "United States", "أمريكا": "United States",
    "روسيا": "Russia", "الصين": "China", "اليابان": "Japan",
    "الهند": "India", "البرازيل": "Brazil", "أستراليا": "Australia",
    "من": "", "في": "", "على": "", "إلى": "", "عن": "",
    "هو": "", "هي": "", "هذا": "", "هذه": "",
    "ما": "", "ماذا": "", "كيف": "", "لماذا": "",
    "لي": "", "لك": "", "أو": "", "مع": "", "بين": "",
    "نوع": "", "أنواع": "", "شكل": "",
}

# ══════════════════════════════════════════════════
# SOURCE 1 — geologyscience.com
# URL pattern: /rocks/{slug}/ or /minerals/{slug}/
# ══════════════════════════════════════════════════
GEOLOGY_SLUGS = {
    # Igneous rocks
    "basalt": ("rocks", "basalt"), "granite": ("rocks", "granite"),
    "rhyolite": ("rocks", "rhyolite"), "andesite": ("rocks", "andesite"),
    "gabbro": ("rocks", "gabbro"), "obsidian": ("rocks", "obsidian"),
    "pumice": ("rocks", "pumice"), "tuff": ("rocks", "tuff"),
    "peridotite": ("rocks", "peridotite"), "diabase": ("rocks", "diabase"),
    "diorite": ("rocks", "diorite"), "syenite": ("rocks", "syenite"),
    "pegmatite": ("rocks", "pegmatite"), "scoria": ("rocks", "scoria"),
    # Sedimentary
    "limestone": ("rocks", "limestone"), "sandstone": ("rocks", "sandstone"),
    "shale": ("rocks", "shale"), "conglomerate": ("rocks", "conglomerate"),
    "breccia": ("rocks", "breccia"), "dolomite": ("rocks", "dolomite"),
    "chalk": ("rocks", "chalk"), "flint": ("rocks", "flint"),
    "chert": ("rocks", "chert"), "marl": ("rocks", "marl"),
    "coal": ("rocks", "coal"), "travertine": ("rocks", "travertine"),
    "mudstone": ("rocks", "mudstone"), "turbidite": ("rocks", "turbidite"),
    # Metamorphic
    "marble": ("rocks", "marble"), "quartzite": ("rocks", "quartzite"),
    "schist": ("rocks", "schist"), "gneiss": ("rocks", "gneiss"),
    "phyllite": ("rocks", "phyllite"), "hornfels": ("rocks", "hornfels"),
    "slate": ("rocks", "slate"), "migmatite": ("rocks", "migmatite"),
    "eclogite": ("rocks", "eclogite"),
    # Minerals
    "quartz": ("minerals", "quartz"), "calcite": ("minerals", "calcite"),
    "feldspar": ("minerals", "feldspar"), "orthoclase": ("minerals", "orthoclase"),
    "plagioclase": ("minerals", "plagioclase"), "muscovite": ("minerals", "muscovite"),
    "biotite": ("minerals", "biotite"), "mica": ("minerals", "mica"),
    "olivine": ("minerals", "olivine"), "pyroxene": ("minerals", "pyroxene"),
    "hornblende": ("minerals", "hornblende"), "amphibole": ("minerals", "amphibole"),
    "garnet": ("minerals", "garnet"), "pyrite": ("minerals", "pyrite"),
    "magnetite": ("minerals", "magnetite"), "hematite": ("minerals", "hematite"),
    "malachite": ("minerals", "malachite"), "azurite": ("minerals", "azurite"),
    "galena": ("minerals", "galena"), "sphalerite": ("minerals", "sphalerite"),
    "chalcopyrite": ("minerals", "chalcopyrite"), "gypsum": ("minerals", "gypsum"),
    "halite": ("minerals", "halite"), "fluorite": ("minerals", "fluorite"),
    "tourmaline": ("minerals", "tourmaline"), "zircon": ("minerals", "zircon"),
    "apatite": ("minerals", "apatite"), "talc": ("minerals", "talc"),
    "chlorite": ("minerals", "chlorite"), "epidote": ("minerals", "epidote"),
    "goethite": ("minerals", "goethite"), "limonite": ("minerals", "limonite"),
    "barite": ("minerals", "barite"), "sulfur": ("minerals", "sulfur"),
    "diamond": ("minerals", "diamond"), "ruby": ("minerals", "ruby"),
    "sapphire": ("minerals", "sapphire"), "emerald": ("minerals", "emerald"),
    "amethyst": ("minerals", "amethyst"), "topaz": ("minerals", "topaz"),
    "opal": ("minerals", "opal"), "jade": ("minerals", "jade"),
    "lapis": ("minerals", "lapis-lazuli"), "onyx": ("minerals", "onyx"),
    "pearl": ("minerals", "pearl"), "turquoise": ("minerals", "turquoise"),
    "gold": ("minerals", "gold"), "silver": ("minerals", "silver"),
    "copper": ("minerals", "copper"),
}

# ══════════════════════════════════════════════════
# SOURCE 2 — geologyistheway.com  ← NEW
# URL patterns discovered from search results:
#   minerals  → /minerals/{slug}/
#   igneous   → /igneous/{slug}/
#   sediment. → /sedimentary/{slug}/
#   metamorph → /metamorphic/{slug}/   (inferred)
# Special cases noted: dolomite mineral → "dolomitemineral",
#   garnet → "garnet-group", amphibole → "amphibole-group"
# ══════════════════════════════════════════════════
GEOLOGYWAY_SLUGS = {
    # Minerals — /minerals/{slug}/
    "quartz":       ("minerals", "quartz"),
    "calcite":      ("minerals", "calcite"),
    "dolomite":     ("minerals", "dolomitemineral"),   # special slug
    "olivine":      ("minerals", "olivine"),
    "garnet":       ("minerals", "garnet-group"),      # special slug
    "amphibole":    ("minerals", "amphibole-group"),   # special slug
    "hornblende":   ("minerals", "hornblende"),
    "pyroxene":     ("minerals", "pyroxene-group"),
    "feldspar":     ("minerals", "feldspar"),
    "plagioclase":  ("minerals", "plagioclase"),
    "orthoclase":   ("minerals", "orthoclase"),
    "muscovite":    ("minerals", "muscovite"),
    "biotite":      ("minerals", "biotite"),
    "mica":         ("minerals", "mica"),
    "chlorite":     ("minerals", "chlorite"),
    "epidote":      ("minerals", "epidote"),
    "tourmaline":   ("minerals", "tourmaline"),
    "zircon":       ("minerals", "zircon"),
    "apatite":      ("minerals", "apatite"),
    "talc":         ("minerals", "talc"),
    "pyrite":       ("minerals", "pyrite"),
    "magnetite":    ("minerals", "magnetite"),
    "hematite":     ("minerals", "hematite"),
    "goethite":     ("minerals", "goethite"),
    "limonite":     ("minerals", "limonite"),
    "malachite":    ("minerals", "malachite"),
    "azurite":      ("minerals", "azurite"),
    "galena":       ("minerals", "galena"),
    "sphalerite":   ("minerals", "sphalerite"),
    "chalcopyrite": ("minerals", "chalcopyrite"),
    "gypsum":       ("minerals", "gypsum"),
    "halite":       ("minerals", "halite"),
    "fluorite":     ("minerals", "fluorite"),
    "barite":       ("minerals", "barite"),
    "sulfur":       ("minerals", "sulfur"),
    "gold":         ("minerals", "gold"),
    "silver":       ("minerals", "silver"),
    "copper":       ("minerals", "copper"),
    "diamond":      ("minerals", "diamond"),
    "ruby":         ("minerals", "ruby"),
    "sapphire":     ("minerals", "sapphire"),
    "emerald":      ("minerals", "emerald"),
    "amethyst":     ("minerals", "amethyst"),
    "topaz":        ("minerals", "topaz"),
    "opal":         ("minerals", "opal"),
    "jade":         ("minerals", "jade"),
    "turquoise":    ("minerals", "turquoise"),
    # Igneous rocks — /igneous/{slug}/
    "granite":      ("igneous", "granite"),
    "rhyolite":     ("igneous", "rhyolite"),
    "basalt":       ("igneous", "basalt"),
    "andesite":     ("igneous", "andesite"),
    "gabbro":       ("igneous", "gabbro"),
    "obsidian":     ("igneous", "obsidian"),
    "pumice":       ("igneous", "pumice"),
    "tuff":         ("igneous", "tuff"),
    "peridotite":   ("igneous", "peridotite"),
    "diorite":      ("igneous", "diorite"),
    "syenite":      ("igneous", "syenite"),
    "pegmatite":    ("igneous", "pegmatite"),
    "scoria":       ("igneous", "scoria"),
    "diabase":      ("igneous", "diabase"),
    "dunite":       ("igneous", "dunite"),
    "kimberlite":   ("igneous", "kimberlite"),
    "phonolite":    ("igneous", "phonolite"),
    "trachyte":     ("igneous", "trachyte"),
    # Sedimentary rocks — /sedimentary/{slug}/
    "limestone":    ("sedimentary", "limestone"),
    "sandstone":    ("sedimentary", "sandstone"),
    "shale":        ("sedimentary", "shale"),
    "conglomerate": ("sedimentary", "conglomerate"),
    "breccia":      ("sedimentary", "breccia"),
    "chalk":        ("sedimentary", "chalk"),
    "flint":        ("sedimentary", "flint"),
    "chert":        ("sedimentary", "chert"),
    "marl":         ("sedimentary", "marl"),
    "coal":         ("sedimentary", "coal"),
    "travertine":   ("sedimentary", "travertine"),
    "mudstone":     ("sedimentary", "mudstone"),
    "turbidite":    ("sedimentary", "turbidite"),
    "dolostone":    ("sedimentary", "dolostone"),
    "evaporite":    ("sedimentary", "evaporite"),
    "phosphorite":  ("sedimentary", "phosphorite"),
    # Metamorphic rocks — /metamorphic/{slug}/
    "marble":       ("metamorphic", "marble"),
    "quartzite":    ("metamorphic", "quartzite"),
    "schist":       ("metamorphic", "schist"),
    "gneiss":       ("metamorphic", "gneiss"),
    "phyllite":     ("metamorphic", "phyllite"),
    "hornfels":     ("metamorphic", "hornfels"),
    "slate":        ("metamorphic", "slate"),
    "migmatite":    ("metamorphic", "migmatite"),
    "eclogite":     ("metamorphic", "eclogite"),
    "blueschist":   ("metamorphic", "blueschist"),
    "amphibolite":  ("metamorphic", "amphibolite"),
    "granulite":    ("metamorphic", "granulite"),
    "serpentinite": ("metamorphic", "serpentinite"),
    "skarn":        ("metamorphic", "skarn"),
    "mylonite":     ("metamorphic", "mylonite"),
    "cataclasite":  ("metamorphic", "cataclasite"),
}

SCRAPE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
}

BAD_KEYWORDS = ["globe", "locator", "orthographic", "Flag_of", "Coat_of",
                 "emblem", "seal", "banner", "logo", "icon", "portrait",
                 "newspaper", "magazine", "article", "stamp", "map-", "-map"]

# ─────────────────── File extraction ───────────────────
def extract_text_from_pdf(file_bytes):
    try:
        import fitz
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        return "".join(page.get_text() for page in doc)[:8000]
    except Exception as e:
        return f"[שגיאה בקריאת PDF: {e}]"

def extract_text_from_docx(file_bytes):
    try:
        from docx import Document
        doc = Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())[:8000]
    except Exception as e:
        return f"[שגיאה בקריאת Word: {e}]"

def extract_text_from_xlsx(file_bytes):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"\n--- גיליון: {sheet} ---\n"
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    text += row_text + "\n"
        return text[:8000]
    except Exception as e:
        return f"[שגיאה בקריאת Excel: {e}]"

def extract_text_from_file(file_bytes, filename, mimetype):
    fname = filename.lower()
    if fname.endswith(".pdf") or "pdf" in mimetype:
        return extract_text_from_pdf(file_bytes), "PDF"
    elif fname.endswith(".docx") or "word" in mimetype:
        return extract_text_from_docx(file_bytes), "Word"
    elif fname.endswith(".xlsx") or "excel" in mimetype or "spreadsheet" in mimetype:
        return extract_text_from_xlsx(file_bytes), "Excel"
    elif fname.endswith(".txt") or "text/plain" in mimetype:
        return file_bytes.decode("utf-8", errors="ignore")[:8000], "טקסט"
    return None, None

# ─────────────────── Translation ───────────────────
def translate_to_english(text):
    result = text
    is_arabic = bool(re.search(r'[\u0600-\u06FF]', text))
    is_hebrew = bool(re.search(r'[\u0590-\u05FF]', text))
    if is_arabic:
        for ar, eng in sorted(AR_TO_EN.items(), key=lambda x: -len(x[0])):
            result = result.replace(ar, f" {eng} " if eng else " ")
        result = re.sub(r'\s+', ' ', result).strip()
        if re.search(r'[\u0600-\u06FF]', result):
            result = re.sub(r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]+', '', result)
            result = re.sub(r'\s+', ' ', result).strip()
    elif is_hebrew:
        for heb, eng in sorted(HE_TO_EN.items(), key=lambda x: -len(x[0])):
            result = result.replace(heb, f" {eng} " if eng else " ")
        result = re.sub(r'\s+', ' ', result).strip()
        if re.search(r'[\u0590-\u05FF]', result):
            result = re.sub(r'[\u0590-\u05FF]+', '', result)
            result = re.sub(r'\s+', ' ', result).strip()
    return result.strip() or "nature"

def is_bad_image(url):
    return any(bad.lower() in url.lower() for bad in BAD_KEYWORDS)

def extract_key_term(english_query):
    """Return the single most specific rock/mineral keyword from the translated query."""
    q = english_query.lower()
    for slug in sorted(GEOLOGY_SLUGS.keys(), key=lambda x: -len(x)):
        if slug in q:
            return slug
    words = [w for w in q.split() if len(w) > 3 and w not in
             ('rock','stone','mineral','type','show','image','photo','picture','about','with')]
    return words[0] if words else q.split()[0]

def _scrape_og_image(url, site_name):
    """
    Generic helper: fetch a page and return its og:image or first large content image.
    Shared by both geology site scrapers.
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        print(f"[{site_name}] BeautifulSoup not installed — skipping")
        return None
    try:
        resp = requests.get(url, timeout=8, headers=SCRAPE_HEADERS)
        print(f"[{site_name}] {url} → HTTP {resp.status_code}")
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, 'html.parser')
        # 1. og:image — best quality hero image
        og = soup.find('meta', property='og:image')
        if og and og.get('content'):
            src = og['content']
            if not is_bad_image(src):
                print(f"[{site_name}] og:image → {src}")
                return src
        # 2. First large content image
        for selector in ['.entry-content img', 'article img', '.post-content img',
                         '.post-body img', 'main img', '.wp-block-image img']:
            for img in soup.select(selector):
                src = img.get('src') or img.get('data-src') or ''
                if not src or src.endswith('.svg') or is_bad_image(src):
                    continue
                try:
                    w = int(img.get('width') or 0)
                    h = int(img.get('height') or 0)
                    if (w and w < 100) or (h and h < 100):
                        continue
                except:
                    pass
                if src.startswith('//'):
                    src = 'https:' + src
                elif src.startswith('/'):
                    src = 'https://' + url.split('/')[2] + src
                print(f"[{site_name}] content img → {src}")
                return src
    except Exception as e:
        print(f"[{site_name}] Error: {e}")
    return None

# ─────────────────── Image source 1: geologyscience.com ───────────────────
def get_geologyscience_image(english_query):
    q = english_query.lower()
    matched = None
    for slug_key in sorted(GEOLOGY_SLUGS.keys(), key=lambda x: -len(x)):
        if slug_key in q:
            matched = GEOLOGY_SLUGS[slug_key]
            break
    if not matched:
        print(f"[GeologySci] No slug match for: '{english_query}'")
        return None
    category, slug = matched
    url = f"https://geologyscience.com/{category}/{slug}/"
    return _scrape_og_image(url, "GeologySci")

# ─────────────────── Image source 2: geologyistheway.com  ← NEW ───────────────────
def get_geologyistheway_image(english_query):
    """
    Scrapes geologyistheway.com using its URL structure:
      minerals   → /minerals/{slug}/
      igneous    → /igneous/{slug}/
      sedimentary→ /sedimentary/{slug}/
      metamorphic→ /metamorphic/{slug}/
    Falls back to None if no slug is mapped or page unavailable.
    """
    q = english_query.lower()
    matched = None
    for slug_key in sorted(GEOLOGYWAY_SLUGS.keys(), key=lambda x: -len(x)):
        if slug_key in q:
            matched = GEOLOGYWAY_SLUGS[slug_key]
            break
    if not matched:
        print(f"[GeologyWay] No slug match for: '{english_query}'")
        return None
    category, slug = matched
    url = f"https://geologyistheway.com/{category}/{slug}/"
    return _scrape_og_image(url, "GeologyWay")

# ─────────────────── Image source 3: Wikipedia pageimages ───────────────────
def get_wikipedia_image(query):
    try:
        api_url = "https://en.wikipedia.org/w/api.php"
        resp = requests.get(api_url, params={
            "action": "query", "generator": "search",
            "gsrsearch": query, "gsrlimit": 5,
            "prop": "pageimages", "pithumbsize": 1200,
            "pilimit": 5, "format": "json"
        }, timeout=8, headers={"User-Agent": "SmartTeacher/1.0"}).json()
        pages = resp.get("query", {}).get("pages", {})
        for page in sorted(pages.values(), key=lambda x: x.get('index', 999)):
            thumb = page.get("thumbnail", {})
            if thumb and thumb.get("source"):
                img_url = re.sub(r'/\d+px-', '/1200px-', thumb["source"])
                if not is_bad_image(img_url):
                    print(f"[Wikipedia] → {img_url}")
                    return img_url
    except Exception as e:
        print(f"[Wikipedia] Error: {e}")
    return None

# ─────────────────── Image source 4: Wikimedia Commons ───────────────────
def get_commons_image(query):
    try:
        resp = requests.get("https://commons.wikimedia.org/w/api.php", params={
            "action": "query", "generator": "search",
            "gsrsearch": query, "gsrnamespace": 6, "gsrlimit": 8,
            "prop": "imageinfo", "iiprop": "url|size", "iiurlwidth": 1200,
            "format": "json"
        }, timeout=6, headers={"User-Agent": "SmartTeacher/1.0"}).json()
        best, best_size = None, 0
        for page in resp.get("query", {}).get("pages", {}).values():
            info = page.get("imageinfo", [{}])[0]
            url  = info.get("thumburl") or info.get("url", "")
            size = info.get("width", 0)
            if url and not is_bad_image(url) and size > best_size:
                best, best_size = url, size
        if best:
            print(f"[Commons] → {best}")
        return best
    except Exception as e:
        print(f"[Commons] Error: {e}")
    return None

# ─────────────────── Master image builder ───────────────────
def build_image_url(user_input):
    is_map = any(word in user_input for word in MAP_WORDS)
    english_query = translate_to_english(user_input)
    key_term = extract_key_term(english_query)
    print(f"[Image] '{user_input}' → '{english_query}' → key='{key_term}'")

    if is_map:
        return (get_wikipedia_image(f"{english_query} map geography") or
                get_commons_image(f"{english_query} map") or
                f"https://picsum.photos/seed/{abs(hash(english_query)) % 1000}/1024/768")

    # Priority 1 — geologyscience.com
    img = get_geologyscience_image(english_query)
    if img:
        return img

    # Priority 2 — geologyistheway.com  ← NEW
    img = get_geologyistheway_image(english_query)
    if img:
        return img

    # Priority 3 — Wikipedia
    img = get_wikipedia_image(f"{key_term} geology")
    if img:
        return img

    # Priority 4 — Wikimedia Commons
    img = get_commons_image(f"{key_term} geology specimen")
    if img:
        return img

    # Fallback
    seed = abs(hash(english_query)) % 1000
    return f"https://picsum.photos/seed/{seed}/1024/768"

# ─────────────────── YouTube video search ───────────────────
def get_youtube_video(query, lang='he'):
    key = YT_API_KEY or API_KEY
    if not key:
        print("[YouTube] ERROR: No API key available")
        return None
    key_term = extract_key_term(query)
    search_query = f"{key_term} geology"
    print(f"[YouTube] Searching: '{search_query}' (lang={lang})")
    for category_id in ["27", "28", ""]:
        params = {
            "part": "snippet", "q": search_query,
            "type": "video", "maxResults": 10,
            "safeSearch": "strict", "videoEmbeddable": "true", "key": key,
        }
        if category_id:
            params["videoCategoryId"] = category_id
        try:
            resp = requests.get("https://www.googleapis.com/youtube/v3/search",
                                params=params, timeout=8)
            data = resp.json()
            if "error" in data:
                err = data["error"]
                print(f"[YouTube] API ERROR {err.get('code')}: {err.get('message')}")
                return None
            items = data.get("items", [])
            print(f"[YouTube] Cat {category_id or 'any'}: {len(items)} results")
            for item in items:
                video_id = item.get("id", {}).get("videoId")
                title    = item.get("snippet", {}).get("title", "").lower()
                if video_id and (key_term in title or query.split()[0].lower() in title):
                    embed = f"https://www.youtube.com/embed/{video_id}?rel=0&modestbranding=1"
                    print(f"[YouTube] ✓ '{title}' → {embed}")
                    return embed
            for item in items:
                video_id = item.get("id", {}).get("videoId")
                title    = item.get("snippet", {}).get("title", "").lower()
                if video_id:
                    embed = f"https://www.youtube.com/embed/{video_id}?rel=0&modestbranding=1"
                    print(f"[YouTube] ~ fallback '{title}' → {embed}")
                    return embed
        except Exception as e:
            print(f"[YouTube] Exception (cat={category_id}): {e}")
    print("[YouTube] No video found")
    return None

def build_video_url(user_input):
    lang = detect_lang(user_input)
    english_query = translate_to_english(user_input)
    print(f"[Video] query='{english_query}' lang={lang}")
    return get_youtube_video(english_query, lang=lang)

# ─────────────────── System prompt ───────────────────
SYSTEM_PROMPT = """אתה 'המורה החכם' — מורה ומדען מנוסה ומומחה.

חוק מוחלט — פתיחת תשובה:
- NEVER start with: "כפרופסור", "בתור מומחה", "אציג", "אתחיל", "ברצוני", "אשמח להציג"
- התחל תמיד ישירות עם תוכן התשובה בהתאם לשאלה:
  * "מה רואים?" → "בתמונה רואים..."
  * "מה זה X?" → "X הוא..."
  * "למה?" → "הסיבה..."
  * "איך?" → "כדי ל..." / "התהליך..."
  * שאלה קצרה בשיחה ממושכת → תשובה ישירה ללא מבוא כלל
- ברמה אקדמית מתאימה לשאלה, בצורה מנומסת וברורה."""

# ─────────────────── Routes ───────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/chat", methods=["POST"])
def chat():
    user_input  = request.form.get("message", "")
    image_file  = request.files.get("image")
    doc_file    = request.files.get("document")
    history_raw = request.form.get("history", "[]")
    try:
        history = json.loads(history_raw)
    except:
        history = []
    if not user_input and not image_file and not doc_file:
        return jsonify({"reply": "Empty message"}), 400

    doc_text, doc_type = None, None
    if doc_file and doc_file.filename:
        file_bytes = doc_file.read()
        doc_text, doc_type = extract_text_from_file(
            file_bytes, doc_file.filename, doc_file.content_type or "")

    has_uploaded_file = (image_file and image_file.filename) or (doc_file and doc_file.filename)
    wants_video  = not has_uploaded_file and any(w in user_input for w in VIDEO_WORDS)
    wants_visual = not has_uploaded_file and not wants_video and \
                   any(w in user_input for w in MAP_WORDS + IMAGE_WORDS)

    image_url = build_image_url(user_input) if wants_visual else None
    video_url = build_video_url(user_input) if wants_video else None

    gemini_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={API_KEY}"
    headers = {'Content-Type': 'application/json'}

    if doc_text:
        user_question = user_input if user_input else "סכם את הקובץ בצורה מפורטת"
        prompt_text = f"""{SYSTEM_PROMPT}
המשתמש העלה קובץ {doc_type}. ענה על שאלתו על סמך תוכן הקובץ בלבד.
תוכן הקובץ:\n---\n{doc_text}\n---\nשאלה: {user_question}"""
    elif wants_video:
        prompt_text = f"""{SYSTEM_PROMPT}
המשתמש ביקש סרטון — המערכת מציגה אותו אוטומטית.
תפקידך: תאר את הנושא בקצרה (2-3 משפטים) כמבוא לסרטון. אל תזכיר שאינך יכול להציג סרטונים.
שאלה: {user_input}"""
    elif wants_visual:
        prompt_text = f"""{SYSTEM_PROMPT}
המשתמש ביקש תמונה/מפה — המערכת מציגה אותה אוטומטית.
תפקידך: תאר את הנושא בהתאם לשאלה. אל תזכיר שאינך יכול להציג תמונות.
שאלה: {user_input}"""
    else:
        prompt_text = f"{SYSTEM_PROMPT}\nשאלה: {user_input}"

    contents = []
    for msg in history:
        role = "user" if msg['role'] == "user" else "model"
        contents.append({"role": role, "parts": [{"text": msg['text']}]})
    current_parts = [{"text": prompt_text}]
    if image_file and image_file.filename:
        try:
            img_data = base64.b64encode(image_file.read()).decode('utf-8')
            current_parts.append({"inline_data": {"mime_type": image_file.content_type, "data": img_data}})
        except:
            pass
    contents.append({"role": "user", "parts": current_parts})

    try:
        response = requests.post(gemini_url, json={"contents": contents}, headers=headers)
        data = response.json()
        if response.status_code == 200:
            reply = data['candidates'][0]['content']['parts'][0]['text']
        else:
            reply = "שגיאת שרת גוגל."
            print(f"[Gemini] Error {response.status_code}: {data}")
        try:
            label = user_input or (doc_file.filename if doc_file else "תמונה")
            db.execute("INSERT INTO history (user_message, bot_message) VALUES (?, ?)", label, reply)
        except:
            pass
        return jsonify({"reply": reply, "image_url": image_url, "video_url": video_url})
    except Exception as e:
        return jsonify({"reply": f"תקלה: {str(e)}", "image_url": image_url, "video_url": video_url})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host='0.0.0.0', port=port)